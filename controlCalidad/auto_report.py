from openpyxl import load_workbook
import pandas as pd

archivo_excel = pd.read_excel('Informe Tranformadores MLU.xlsx')

tabla_pivote = archivo_excel[['Fecha', 'TERRITORIO', 'CIRCUITO', 'ID_BDI','CODELEME', 'CODIGO_BDI_CIRCUITO', 'PLACA MT COLOCADA', 'Equipo Ruta Id', 'PLACA MT ANTERIOR',
'Matricula MT anterior', 'POTENCIA', 'POTENCIA NOMINAL MODIFICADA', 'FABRICANTE', 'MARCA MODIFICADA', 'OTRA MARCA', 'RELTRANS', 'TENSIÓN SECUNDARIA MODIFICADA (V)',
'TIPINS', 'CONFIRME TIPO DE INSTALACIÓN', 'TIPOFASE','CONFIRME TIPO DE TRANSFORMADOR', 'ESTADO DEL TRANSFORMADOR', 'Longitud del equipo padre', 'Latitud del equipo padre',
'FOTO VERIFICACIÓN FRENTE - LADO 1','FOTO VERIFICACIÓN FRENTE - LADO 2', 'Foto matricula MT','FOTO MT COLOCADA', 'SOPORTE 01','SOPORTE 02', 'SOPORTE FOTOGRÁFICO 03', 'SOPORTE FOTOGRÁFICO 04', 'Nombre Equipo padre',
'Nombre del Usuario', 'Longitud', 'Latitud']]

#Eliminar duplicados equipo ruta id
tabla_pivote = tabla_pivote.drop_duplicates(subset=['Equipo Ruta Id']) 

#rellenar vacias por ceros
tabla_pivote.fillna(0, inplace=True)

# print(tabla_pivote)
tabla_pivote.to_excel('Auto_Reporte.xlsx', sheet_name='CC')

wb = load_workbook('Auto_Reporte.xlsx')
pestaña = wb['CC']

min_colm = wb.active.min_column
max_colm = wb.active.max_column
min_fila = wb.active.min_row
max_fila = wb.active.max_row

#print(min_colm)
#print(max_colm)
#print(min_fila)
#print(max_fila)


#generacion de codigo id_bdi, codeleme o nuevo
pestaña.cell(row=1,column=5).value = 'CODIGO'
for i in range(2,max_fila+1):
        if pestaña.cell(row=i,column=6).value == 0:
           pestaña.cell(row=i,column=6).value = str(10) + str(pestaña.cell(row=i,column=9).value)
        if pestaña.cell(row=i,column=5).value == 0:
                pestaña.cell(row=i,column=5).value = pestaña.cell(row=i,column=6).value

#INSTALACION SUP
pestaña.cell(row=1,column=7).value = 'Instalacion_Superior'

#PLACA MT ANTERIOR
pestaña.cell(row=1,column=10).value = 'Matricula_Antigua'
for i in range(2,max_fila+1):
        if pestaña.cell(row=i,column=11).value != 0:
                pestaña.cell(row=i,column=10).value = pestaña.cell(row=i,column=11).value

#potencia
pestaña.cell(row=1,column=12).value = 'Potencia_Nominal_(kVA)'
for i in range(2,max_fila+1):
        if pestaña.cell(row=i,column=13).value != 0:
                pestaña.cell(row=i,column=12).value = pestaña.cell(row=i,column=13).value

#marca
pestaña.cell(row=1,column=14).value = 'Marca'
for i in range(2,max_fila+1):
        if pestaña.cell(row=i,column=16).value != 0:
                pestaña.cell(row=i,column=15).value = pestaña.cell(row=i,column=16).value
        if pestaña.cell(row=i,column=15).value != 0:
                pestaña.cell(row=i,column=14).value = pestaña.cell(row=i,column=15).value

#tension secundaria
pestaña.cell(row=1,column=17).value = 'Tension_Secundaria_(V)'
for i in range(2,max_fila+1):
        if pestaña.cell(row=i,column=18).value != 0:
                pestaña.cell(row=i,column=17).value = pestaña.cell(row=i,column=18).value

#ubicacion-tipo de instalacion
pestaña.cell(row=1,column=19).value = 'Ubicacion_Trafo'
for i in range(2,max_fila+1):
        if pestaña.cell(row=i,column=20).value != 0:
                pestaña.cell(row=i,column=19).value = pestaña.cell(row=i,column=20).value
        if pestaña.cell(row=i,column=19).value == 'AÉREO':
                pestaña.cell(row=i,column=19).value = 'AEREO'
        if pestaña.cell(row=i,column=19).value == 'SUBTERRÁNEO':
                pestaña.cell(row=i,column=19).value = 'SUBTERRANEO'

#TIPO DE TX
pestaña.cell(row=1,column=21).value = 'TIPOFASE'
for i in range(2,max_fila+1):
        if pestaña.cell(row=i,column=22).value != 0:
                pestaña.cell(row=i,column=21).value = pestaña.cell(row=i,column=22).value
        if pestaña.cell(row=i,column=21).value == 'MONOFÁSICO BIFILAR':
                pestaña.cell(row=i,column=21).value = 'MONOFASICO BIFILAR'
        if pestaña.cell(row=i,column=21).value == 'MONOFÁSICO':
                pestaña.cell(row=i,column=21).value = 'MONOFASICO'
        if pestaña.cell(row=i,column=21).value == 'TRIFÁSICO':
                pestaña.cell(row=i,column=21).value = 'TRIFASICO'

#LONGITUD
pestaña.cell(row=1,column=24).value = 'Longitud (WGS84)'
for i in range(2,max_fila+1):
        if pestaña.cell(row=i,column=24).value == 0:
                pestaña.cell(row=i,column=24).value = pestaña.cell(row=i,column=36).value

#LATITUD
pestaña.cell(row=1,column=25).value = 'Latitud (WGS84)'
for i in range(2,max_fila+1):
        if pestaña.cell(row=i,column=25).value == 0:
                pestaña.cell(row=i,column=25).value = pestaña.cell(row=i,column=37).value

#MOVIENDO, ELIMINANDO Y CREANDO COLUMNAS
for i in range(1,max_fila+1):
        pestaña.cell(row=i,column=6).value = pestaña.cell(row=i,column=7).value
        pestaña.cell(row=i,column=7).value = pestaña.cell(row=i,column=8).value
        pestaña.cell(row=i,column=8).value = pestaña.cell(row=i,column=9).value
        pestaña.cell(row=i,column=9).value = pestaña.cell(row=i,column=10).value
        pestaña.cell(row=i,column=10).value = pestaña.cell(row=i,column=12).value
        pestaña.cell(row=i,column=11).value = pestaña.cell(row=i,column=14).value
        pestaña.cell(row=i,column=12).value = '=SI(L2="240/120";"240";SI(L2="214/124";"214124";SI(L2="213/123";"213123";SI(L2="208/120";"208";SI(L2="220/127";"220";SI(L2="214/123";"214";SI(L2="214/123.6";"214/123.6";SI(L2="228/132";"228132";SI(L2="226/130";"226130";SI(L2="231/133.4";"231";SI(L2="228.6/132";"2286132";SI(L2="ILEGIBLE";"ILEGIBLE";SI(L2="225/130";"225130";SI(L2="231/133";"231133";SI(L2="454/262";"454262";SI(L2="228/131.6";"228";SI(L2="225/129.9";"225";SI(L2="480/277";"480277";SI(L2="480/277.1";"480";SI(L2="451/260";"451260";"PENDIENTE"))))))))))))))))))))'
        pestaña.cell(row=i,column=13).value = pestaña.cell(row=i,column=17).value
        pestaña.cell(row=i,column=14).value = '=SI(N2="AEREO";"AE";SI(N2="SUBTERRANEO";"SB";SI(N2="SUPERFICIE";"SP";0)))'
        pestaña.cell(row=i,column=15).value = pestaña.cell(row=i,column=19).value
        pestaña.cell(row=i,column=16).value = '=SI(P2="MONOFASICO BIFILAR";1;SI(P2="TRIFASICO";3;2))'
        pestaña.cell(row=i,column=17).value = pestaña.cell(row=i,column=21).value
        pestaña.cell(row=i,column=18).value = '=SI(O(R2="EN SERVICIO";R2="MATRI/NVO");"SE";SI(O(R2="DESCONECTADO";R2="MATRI/DESC";R2="MATRI/NVO/DESC");"DE"))'
        pestaña.cell(row=i,column=19).value = pestaña.cell(row=i,column=23).value
        pestaña.cell(row=i,column=20).value = None
        pestaña.cell(row=i,column=21).value = None
        pestaña.cell(row=i,column=22).value = '17'
        pestaña.cell(row=i,column=23).value = 'CENSO II'
        
        pestaña.cell(row=i,column=38).value = pestaña.cell(row=i,column=35).value
        pestaña.cell(row=i,column=37).value = pestaña.cell(row=i,column=34).value
        pestaña.cell(row=i,column=36).value = pestaña.cell(row=i,column=33).value
        pestaña.cell(row=i,column=35).value = pestaña.cell(row=i,column=32).value
        pestaña.cell(row=i,column=34).value = pestaña.cell(row=i,column=31).value
        pestaña.cell(row=i,column=33).value = pestaña.cell(row=i,column=30).value
        pestaña.cell(row=i,column=32).value = pestaña.cell(row=i,column=29).value
        pestaña.cell(row=i,column=31).value = pestaña.cell(row=i,column=28).value
        pestaña.cell(row=i,column=30).value = pestaña.cell(row=i,column=27).value
        pestaña.cell(row=i,column=29).value = pestaña.cell(row=i,column=26).value
        pestaña.cell(row=i,column=28).value = pestaña.cell(row=i,column=25).value
        pestaña.cell(row=i,column=27).value = pestaña.cell(row=i,column=24).value

        pestaña.cell(row=i,column=24).value = None
        pestaña.cell(row=i,column=25).value = '=SI(Y2="URBANO";"U";SI(Y2="RURAL";"R";"ERROR"))'
        pestaña.cell(row=i,column=26).value = None

#nombrando columnas
pestaña.cell(row=1,column=12).value = 'TENS_SEC'
pestaña.cell(row=1,column=14).value = 'UBICACION_TRAFO'
pestaña.cell(row=1,column=16).value = 'Tipo_de_Conexion'
pestaña.cell(row=1,column=18).value = 'Estado_Elemento'
pestaña.cell(row=1,column=20).value = 'COD'
pestaña.cell(row=1,column=21).value = 'Observaciones'
pestaña.cell(row=1,column=22).value = 'Origen_de_los_datos'
pestaña.cell(row=1,column=23).value = 'Origen_de_los_datos'
pestaña.cell(row=1,column=24).value = 'UC_R015'
pestaña.cell(row=1,column=25).value = 'TIPO_AREA'
pestaña.cell(row=1,column=26).value = 'TIPO DE AREA'
pestaña.cell(row=1,column=39).value = 'ESTADO COORDENADAS'

wb.save('Auto_Reporte.xlsx')
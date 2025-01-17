from openpyxl import load_workbook
import pandas as pd

archivoCargue = pd.read_excel('FEHCA_CIRCUITO_DPTO_ACCESO_IMP_CODIGO.xlsx')
archivoCodigo = pd.read_excel('CODIGOS_CIRCUITO.xlsx')

wb = load_workbook('FEHCA_CIRCUITO_DPTO_ACCESO_IMP_CODIGO.xlsx')
datos = wb['DATOS_RUTA_TRABAJO']

condicion_No= input("ACCESO_IMP - 0   CIERRE_CTO - 1  :")

if condicion_No == 0:
    condicion = "ACCESO_IMP"
else:
    condicion = " CIERRE_CTO"
    
seguir = "S"
while (seguir == "S") or (seguir == "s") :
    fecha = "20221012"
    circuito = "none"
    codigo = 0
    departamento = "none"
    NoRuta = "none"
    nombreKml = "none"
    
    circuito = input("CIRCUITO : ")
    NoRuta = input("Numero de Ruta : ")
    # nombreKml = input("Nombre del kml : ")
    for i in archivoCodigo.index:
        if archivoCodigo["CIRCUITO (ID 147)"][i] == circuito:
            codigo = int(archivoCodigo["ID"][i])
            departamento = archivoCodigo["TERRITORIO (ID 145)"][i]
    
    if codigo == 0:
        codigo = input("CODIGO : ")
        departamento = input("TERRITORIO : ")
    
    xcircuito = circuito.replace(' ', '_') #remplazamos
    nombreKml = xcircuito
    
    print('MATRICULACION_' + str(codigo))
    print("MATRICULACIÓN")
    print(str(codigo))
    print(str(xcircuito)+"_"+str(departamento)+"_"+condicion+"_R00"+str(NoRuta))
    print(str(xcircuito)+"_"+str(departamento)+"_"+condicion+"_R00"+str(NoRuta))
    print(str(nombreKml))
    
    datos.cell(row=2,column=1).value = 'MATRICULACION_' + str(codigo) 
    datos.cell(row=2,column=2).value = "MATRICULACIÓN"
    datos.cell(row=2,column=3).value = str(codigo)
    datos.cell(row=2,column=4).value = str(xcircuito)+"_"+str(departamento)+"_"+condicion+"_R00"+str(NoRuta)
    datos.cell(row=2,column=5).value = str(xcircuito)+"_"+str(departamento)+"_"+condicion+"_R00"+str(NoRuta)
    datos.cell(row=2,column=6).value = str(nombreKml)
    
    wb.save(fecha+'_'+xcircuito+'_'+departamento+'_'+condicion+'_'+str(codigo)+'.xlsx')
    
    seguir = input("crear otra ruta : ")
